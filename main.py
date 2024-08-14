from sleeper_wrapper import User, League, Stats
from dotenv import load_dotenv
import os
import json
from getplayers import get_or_generate_players_files 
from fileutil import save_json_file
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Load environment variables from .env file
load_dotenv()

# Get user ID and league ID from environment variables
user_id = os.getenv('USER_ID')
league_id = os.getenv('LEAGUE_ID')

# Create a User instance
user = User(user_id)

# Get user details to fetch the user ID
user_details = user.get_user()
actual_user_id = user_details['user_id']

# Create a League instance
league = League(league_id)

stats_client = Stats()

# Fetch league details
league_details = league.get_league()

if (league_details['status'] != 'complete'):
    print("League " + league_details['name'] + league_details['season'] + " is not complete yet. Check that you configured the correct league ID. To find it, follow instructions here: https://support.sleeper.com/en/articles/4121798-how-do-i-find-my-league-id ")
else:
    scoring_settings = league_details['scoring_settings']
    settings = league_details['settings']
    weeks_in_season = settings['last_scored_leg']
    
    def get_week_stats(week): 
        # Check if the file already exists
        tmp_folder = os.path.join(os.path.dirname(__file__), 'tmp')
        if not os.path.exists(tmp_folder):
            os.makedirs(tmp_folder)
        
        filename = f"stats{league_details['season']}{week}"
        json_file_name = filename + '.json'
        json_file = os.path.join(tmp_folder, json_file_name)
        csv_file_path = os.path.join(tmp_folder, filename + '.csv')
        week_stats = {}
        
        if not os.path.exists(json_file):
            # Get player stats for each week
            week_stats = stats_client.get_week_stats('regular', league_details['season'], week) 
            # Write week_stats to the file
            with open(csv_file_path, 'w') as file:
                for player_id, stats in week_stats.items():
                    file.write(f"{player_id},{','.join(str(stat) for stat in stats)}\n")
            save_json_file(week_stats, json_file_name)
        else: 
            with open(json_file, 'r') as file:
                week_stats = json.load(file)
        return week_stats
    
    # TODO calculate values: 
    # top_qb_week = ""
    # top_qb_week_score = -1
    # top_rb_week = ""
    # top_rb_week_score = -1
    # top_wr_week = ""
    # top_wr_week_score = -1
    # top_te_week = ""
    # top_te_week_score = -1
    def get_player_scores():
        player_scores = {}
    
        # only get scores for the weeks that were a part of the season
        for week in range(1, weeks_in_season + 1):
            week_stats = get_week_stats(week)
            
            for player_id, stats in week_stats.items():
                player_score = 0
                def calculate_score_for_this_week(player_score, valid_stats):
                    for stat in valid_stats:
                        # find the stat in the scoring settings
                        if (scoring_settings.get(stat, False)):
                            # if the stat is found, add the value to the player's score
                            points_for_each = scoring_settings[stat]
                            times_it_occured = valid_stats[stat]
                            # if times_it_occured < 0:
                            #     print('tst')
                            # if points_for_each < 0:
                            #     print('tst')
                            player_score += points_for_each * times_it_occured
                    return player_score                    
                
                if "TEAM" in player_id:
                #     # it's a defense, handle scoring differently
                #     #for stat in stats:
                #         #if stat is not a defensive stat, remove it
                #     #valid_stats=
                #     #calculate_score_for_this_week(valid_stats)
                    # print('t')
                    player_score=0
                else: 
                    # it's a player
                    # if player_id == '4046':
                    #     print('t')
                    player_score = calculate_score_for_this_week(player_score, stats)
                    player_scores.setdefault(player_id, 0) # create a new key if it doesn't exist
                    player_scores[player_id] += player_score       
        return player_scores
    
    # assign a cost to each player
    def get_player_costs(players, sorted_player_scores):
        players_by_position = {}
        player_costs = {}
        # for each player score, add the player to the list for that position
        for player_score in sorted_player_scores: 
            players_by_position.setdefault(players[player_score[0]]['position'], []).append(players[player_score[0]]['player_id']) # full_name can be used to debug by player names in place of player_id
        def set_position_salaries(players_by_position, position): 
            SALARIES = os.getenv(position + '_SALARIES').split(',')
            for player in players_by_position[position]:
                player_costs[player] = SALARIES[0]
                if len(SALARIES) != 1: 
                    # the last salary is default for everyone else
                    SALARIES.pop(0) 
            return player_costs
        
        set_position_salaries(players_by_position, "QB")
        set_position_salaries(players_by_position, "RB")
        set_position_salaries(players_by_position, "WR")
        set_position_salaries(players_by_position, "TE")
        set_position_salaries(players_by_position, "DEF")
        return player_costs
    
        
    players = get_or_generate_players_files()
    player_scores = get_player_scores()
    
    # Sort the dictionary by values in descending order and convert to a list of tuples
    sorted_player_scores = sorted(player_scores.items(), key=lambda item: item[1], reverse=True)
    
    # for each sorted player, get the position from the players dictionary and pop the cost off the position list
    player_costs = get_player_costs(players, sorted_player_scores)
    
    # get all rosters
    rosters = league.get_rosters()
    league_users = league.get_users()

    # Create a dictionary to store player ranks
    position_counters = {}
    player_ranks = {}

    for player_id, _ in sorted_player_scores:
        position = players[player_id]['position']
        if position not in position_counters:
            position_counters[position] = 1
        else:
            position_counters[position] += 1
        player_ranks[player_id] = f"{position}{position_counters[position]}"

    def write_workbook():
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write the header row
        header = []
        for roster in rosters:
            owner_id = roster['owner_id']
            # Search for the team name or display name in league_users
            owner_info = next((user for user in league_users if user['user_id'] == owner_id), None)
            if owner_info:
                owner_team_name = owner_info['metadata'].get('team_name')
                display_name = owner_info.get('display_name', 'Unknown')
                if owner_team_name:
                    header.append(owner_team_name)
                    header.append(display_name)
                    header.append("")
                else:
                    header.append(display_name)
                    header.append("")
                    header.append("")
            else:
                header.append('Unknown')
        ws.append(header)

        # Write the second row
        second_row = []
        for _ in rosters:
            second_row.extend(['Player', 'Rank', 'Cost'])
        ws.append(second_row)

        # Determine the maximum number of players in any roster
        max_players = max(len(roster['players']) for roster in rosters)

        # Write each player's ID, rank, and cost for each roster
        for i in range(max_players):
            row = []
            for roster in rosters:
                if i < len(roster['players']):
                    player_id = roster['players'][i]
                    player_cost = float(player_costs.get(player_id, ''))
                    player = players[player_id]
                    # Use team if full_name doesn't exist, since it's a defense
                    player_name = player.get('full_name', player.get('team', 'Unknown'))
                    player_rank = player_ranks.get(player_id, '')
                    row.extend([player_name, player_rank, player_cost])
                else:
                    row.extend(['', '', ''])  # Fill with blanks if no more players in this roster
            ws.append(row)

        # Apply vertical borders
        thin_border = Border(left=Side(style='thin'))

        for col in range(1, len(rosters) * 3 + 1, 3):
            for row in range(1, max_players + 3):  # +3 to include header and second row
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

        # Create a black border style for the bottom
        black_bottom_border = Border(bottom=Side(style='thin', color='000000'))
        # Apply the black bottom border to each cell in row 2
        for col in range(1, ws.max_column + 1):
            ws.cell(row=2, column=col).border = black_bottom_border

        # Draw a red line across the bottom of row 22
        red_border = Border(bottom=Side(style='thin', color='FF0000'))
        for col in range(1, len(rosters) * 3 + 1):
            cell = ws.cell(row=22, column=col)
            cell.border = red_border
            
        # Draw a black line across the bottom row
        bottom_row = ws.max_row + 1
        black_border = Border(bottom=Side(style='thin', color='000000'))
        for col in range(1, len(rosters) * 3 + 1):
            ws.cell(row=bottom_row, column=col).border = black_border

        # Insert "Total:" label and SUM formula
        for col in range(3, len(rosters) * 3 + 1, 3):
            # total_label_cell = ws.cell(row=bottom_row + 1, column=col - 1, value="Total:")
            sum_formula = f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{max_players + 2})"
            total_cell = ws.cell(row=bottom_row + 1, column=col, value=sum_formula)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Add specified lines in column A and "dummy val" in column B
        season_payout_keys = [
            "$10 1st place regular season",
            "$10 2nd place regular season",
            "$10 Top qb week",
            "$10 Top rb week",
            "$10 Top wr week",
            "$10 Top te week",
            "$20 first place championship",
            "$10 second place championship",
            "$10 third place championship"
        ]

        # Find the first empty row in column A
        first_empty_row = ws.max_row + 2  # +2 to add a blank line

        for i, line in enumerate(season_payout_keys, start=first_empty_row):
            ws.cell(row=i, column=1, value=line)
            ws.cell(row=i, column=2, value="dummy val")

            
        # Save the workbook
        wb.save('roster_costs.xlsx')

    write_workbook()
    

    print('end')